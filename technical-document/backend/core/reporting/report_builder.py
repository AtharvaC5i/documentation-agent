import os
import json
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, Any
from dotenv import load_dotenv

load_dotenv()

STORAGE_DIR = os.getenv("STORAGE_DIR", os.path.join("..", "storage"))


def _load_json(path: Path) -> dict:
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return {}


def build_report(project_id: str, metadata: Dict[str, Any]) -> Dict[str, Any]:
    """
    Assembles a full run report from all persisted phase outputs.
    No new data collection — reads what already exists on disk.
    """
    base = Path(STORAGE_DIR) / "projects" / project_id

    review_state = _load_json(base / "review_state.json")
    publish_log  = _load_json(base / "publish_log.json")
    assembly_meta_path = base / "assembly_meta.json"
    assembly_meta = _load_json(assembly_meta_path)

    # ── Per-section detail ────────────────────────────────────
    section_rows = []
    total_words  = 0
    scores       = []

    for name, entry in sorted(
        review_state.items(), key=lambda x: x[1].get("order", 0)
    ):
        content  = entry.get("edited_content") or entry.get("original_content", "")
        wc       = len(content.split())
        score    = entry.get("quality_score", 0.0)
        edited   = entry.get("edited_content") is not None
        total_words += wc
        if score:
            scores.append(score)

        section_rows.append({
            "name":          name,
            "word_count":    wc,
            "quality_score": round(score, 2),
            "status":        entry.get("status", "pending"),
            "regenerated":   entry.get("regenerated", False),
            "edited":        edited,
            "note":          entry.get("note", ""),
            "reviewed_at":   entry.get("reviewed_at", ""),
            "order":         entry.get("order", 0),
        })

    avg_score    = round(sum(scores) / len(scores), 2) if scores else 0.0
    score_green  = sum(1 for s in scores if s >= 0.7)
    score_amber  = sum(1 for s in scores if 0.5 <= s < 0.7)
    score_red    = sum(1 for s in scores if s < 0.5)

    total_sections = len(section_rows)
    approved = sum(1 for r in section_rows if r["status"] == "approved")
    rejected = sum(1 for r in section_rows if r["status"] == "rejected")
    edited   = sum(1 for r in section_rows if r["edited"])
    regen    = sum(1 for r in section_rows if r["regenerated"])

    return {
        "project_id":       project_id,
        "project_name":     metadata.get("project_name", "Untitled"),
        "client_name":      metadata.get("client_name", ""),
        "generated_at":     assembly_meta.get("generated_at", ""),
        "report_created_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_sections":  total_sections,
            "approved":        approved,
            "rejected":        rejected,
            "edited":          edited,
            "regenerated":     regen,
            "total_words":     total_words,
            "page_estimate":   max(1, round(total_words / 300)),
            "avg_quality":     avg_score,
            "quality_green":   score_green,
            "quality_amber":   score_amber,
            "quality_red":     score_red,
        },
        "publish": {
            "published":   publish_log.get("published", False),
            "deliveries":  publish_log.get("deliveries", []),
        },
        "sections": section_rows,
    }


def export_report_xlsx(project_id: str, metadata: Dict[str, Any]) -> Path:
    """Generate a formatted .xlsx report and return its path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required for xlsx export. pip install openpyxl")

    report   = build_report(project_id, metadata)
    summary  = report["summary"]
    sections = report["sections"]

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="0D1B3E")
    label_font     = Font(bold=True, color="0D1B3E", size=10)
    center_align   = Alignment(horizontal="center", vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")

    ws1.append(["DocAgent — Run Report"])
    ws1["A1"].font = Font(bold=True, size=14, color="0D1B3E")
    ws1.append([f"Project: {report['project_name']}"])
    ws1.append([f"Client:  {report['client_name']}"])
    ws1.append([f"Report generated: {report['report_created_at'][:19].replace('T', ' ')} UTC"])
    ws1.append([])

    summary_headers = ["Metric", "Value"]
    ws1.append(summary_headers)
    for cell in ws1[ws1.max_row]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    summary_rows = [
        ("Total Sections",      summary["total_sections"]),
        ("Approved",            summary["approved"]),
        ("Rejected",            summary["rejected"]),
        ("Edited by Reviewer",  summary["edited"]),
        ("Auto-Regenerated",    summary["regenerated"]),
        ("Total Word Count",    f"{summary['total_words']:,}"),
        ("Estimated Pages",     summary["page_estimate"]),
        ("Avg Quality Score",   f"{summary['avg_quality']:.0%}"),
        ("High Quality (≥70%)", summary["quality_green"]),
        ("Medium (50–69%)",     summary["quality_amber"]),
        ("Low Quality (<50%)",  summary["quality_red"]),
        ("Published",           "Yes" if report["publish"]["published"] else "No"),
    ]
    for row in summary_rows:
        ws1.append(row)
        ws1.cell(ws1.max_row, 1).font = label_font

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Per-Section Detail ──────────────────────────
    ws2 = wb.create_sheet("Section Detail")

    detail_headers = [
        "#", "Section Name", "Words", "Quality Score",
        "Status", "Regenerated", "Edited", "Reviewer Note", "Reviewed At"
    ]
    ws2.append(detail_headers)
    for cell in ws2[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    green_fill = PatternFill("solid", fgColor="D1FAE5")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")
    stripe     = PatternFill("solid", fgColor="F8F9FF")

    for idx, sec in enumerate(sections, 1):
        score = sec["quality_score"]
        row   = [
            idx,
            sec["name"],
            sec["word_count"],
            f"{score:.0%}",
            sec["status"].capitalize(),
            "Yes" if sec["regenerated"] else "No",
            "Yes" if sec["edited"]      else "No",
            sec["note"]        or "",
            sec["reviewed_at"][:19].replace("T", " ") + " UTC" if sec["reviewed_at"] else "",
        ]
        ws2.append(row)
        row_num = ws2.max_row

        # Quality score colour
        score_cell = ws2.cell(row_num, 4)
        if score >= 0.7:
            score_cell.fill = green_fill
        elif score >= 0.5:
            score_cell.fill = amber_fill
        else:
            score_cell.fill = red_fill

        # Alternating row stripe
        if idx % 2 == 0:
            for col in range(1, len(detail_headers) + 1):
                c = ws2.cell(row_num, col)
                if c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                    c.fill = stripe

        for cell in ws2[row_num]:
            cell.alignment = left_align

    col_widths = [4, 36, 8, 14, 12, 13, 8, 30, 22]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = width

    ws2.freeze_panes = "A2"

    out_path = Path(STORAGE_DIR) / "projects" / project_id / "run_report.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path